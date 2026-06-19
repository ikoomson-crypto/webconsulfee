from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, make_response
import os
import sys
import platform
from datetime import datetime, date
from werkzeug.utils import secure_filename
import csv
import json
from io import BytesIO, StringIO
import openpyxl
from openpyxl import Workbook
import time
import base64
import zipfile
import tempfile
import psycopg2
from psycopg2.extras import RealDictCursor

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# ============ ENVIRONMENT CONFIGURATION ============
IS_RENDER = os.environ.get('RENDER', False) or os.environ.get('RENDER_APP_NAME', False)
IS_WINDOWS = platform.system() == 'Windows'
DATABASE_URL = os.environ.get('DATABASE_URL')

print(f"🖥️ Running on: {'Render' if IS_RENDER else 'Local'}")
print(f"💻 Platform: {platform.system()}")
print(f"🗄️ Using PostgreSQL: {bool(DATABASE_URL)}")

if IS_RENDER:
    DATA_DIR = '/opt/render/project/src/data'
    os.makedirs(DATA_DIR, exist_ok=True)
    app.config['UPLOAD_FOLDER'] = os.path.join(DATA_DIR, 'uploads')
    app.config['EXPORT_FOLDER'] = os.path.join(DATA_DIR, 'exports')
    app.config['LOGO_UPLOAD_FOLDER'] = os.path.join(DATA_DIR, 'static/uploads/logos')
else:
    app.config['UPLOAD_FOLDER'] = 'uploads'
    app.config['EXPORT_FOLDER'] = 'exports'
    app.config['LOGO_UPLOAD_FOLDER'] = 'static/uploads/logos'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)
os.makedirs(app.config['LOGO_UPLOAD_FOLDER'], exist_ok=True)

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg'}


def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ============ DATABASE CONNECTION ============
def get_db():
    if DATABASE_URL:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    else:
        import sqlite3
        conn = sqlite3.connect('payment_system.db')
        conn.row_factory = sqlite3.Row
        return conn


def init_db():
    if DATABASE_URL:
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS suppliers (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                role TEXT,
                address TEXT,
                telephone TEXT,
                bank_name TEXT,
                account_number TEXT,
                account_name TEXT,
                swift_code TEXT,
                bank_address TEXT,
                id_type TEXT,
                id_number TEXT,
                street_address TEXT,
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS services (
                id SERIAL PRIMARY KEY,
                supplier_id INTEGER REFERENCES suppliers(id),
                service_description TEXT NOT NULL,
                total REAL DEFAULT 0,
                status TEXT DEFAULT 'Pending',
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS payments (
                id SERIAL PRIMARY KEY,
                supplier_id INTEGER REFERENCES suppliers(id),
                service_id INTEGER REFERENCES services(id),
                amount REAL,
                payment_method TEXT,
                reference_number TEXT,
                payment_date DATE,
                notes TEXT,
                payment_type TEXT DEFAULT 'Earning',
                status TEXT DEFAULT 'Completed',
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS company_settings (
                id SERIAL PRIMARY KEY,
                company_name TEXT NOT NULL,
                company_address TEXT,
                company_phone TEXT,
                company_email TEXT,
                company_website TEXT,
                company_registration TEXT,
                currency_symbol TEXT DEFAULT '$',
                currency_code TEXT DEFAULT 'USD',
                tax_rate REAL DEFAULT 0,
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute("SELECT id FROM company_settings LIMIT 1")
        existing = cursor.fetchone()
        if not existing:
            cursor.execute('''
                INSERT INTO company_settings (
                    company_name, company_address, company_phone, 
                    company_email, company_website, company_registration,
                    currency_symbol, currency_code, tax_rate
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                'My Company',
                '123 Business Street, City, Country',
                '+1 234 567 8900',
                'info@mycompany.com',
                'www.mycompany.com',
                'REG-2024-001',
                '$',
                'USD',
                0.0
            ))

        conn.commit()
        conn.close()
    else:
        import sqlite3
        conn = sqlite3.connect('payment_system.db')
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                role TEXT,
                address TEXT,
                telephone TEXT,
                bank_name TEXT,
                account_number TEXT,
                account_name TEXT,
                swift_code TEXT,
                bank_address TEXT,
                id_type TEXT,
                id_number TEXT,
                street_address TEXT,
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS services (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_id INTEGER,
                service_description TEXT NOT NULL,
                total REAL DEFAULT 0,
                status TEXT DEFAULT 'Pending',
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_id INTEGER,
                service_id INTEGER,
                amount REAL,
                payment_method TEXT,
                reference_number TEXT,
                payment_date DATE,
                notes TEXT,
                payment_type TEXT DEFAULT 'Earning',
                status TEXT DEFAULT 'Completed',
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
                FOREIGN KEY (service_id) REFERENCES services(id)
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS company_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name TEXT NOT NULL,
                company_address TEXT,
                company_phone TEXT,
                company_email TEXT,
                company_website TEXT,
                company_registration TEXT,
                currency_symbol TEXT DEFAULT '$',
                currency_code TEXT DEFAULT 'USD',
                tax_rate REAL DEFAULT 0,
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute("SELECT id FROM company_settings LIMIT 1")
        existing = cursor.fetchone()
        if not existing:
            cursor.execute('''
                INSERT INTO company_settings (
                    company_name, company_address, company_phone, 
                    company_email, company_website, company_registration,
                    currency_symbol, currency_code, tax_rate
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                'My Company',
                '123 Business Street, City, Country',
                '+1 234 567 8900',
                'info@mycompany.com',
                'www.mycompany.com',
                'REG-2024-001',
                '$',
                'USD',
                0.0
            ))

        conn.commit()
        conn.close()


init_db()


# ============ PDF LIBRARY SETUP ============
PDF_LIBRARY = None
pdf_config = None

if IS_RENDER:
    print("📦 Using WeasyPrint for PDF generation")
    try:
        from weasyprint import HTML
        PDF_LIBRARY = 'weasyprint'
        print("✅ WeasyPrint loaded successfully")
    except ImportError as e:
        print(f"⚠️ WeasyPrint import error: {e}")
        PDF_LIBRARY = None
else:
    print("📦 Using pdfkit for PDF generation")
    try:
        import pdfkit
        WKHTMLTOPDF_PATH = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
        if os.path.exists(WKHTMLTOPDF_PATH):
            pdf_config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)
            print(f"✅ wkhtmltopdf found at: {WKHTMLTOPDF_PATH}")
        else:
            alt_path = r'C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe'
            if os.path.exists(alt_path):
                pdf_config = pdfkit.configuration(wkhtmltopdf=alt_path)
                print(f"✅ wkhtmltopdf found at: {alt_path}")
            else:
                print(f"⚠️ wkhtmltopdf not found")
                pdf_config = None
        PDF_LIBRARY = 'pdfkit'
    except ImportError as e:
        print(f"⚠️ pdfkit import error: {e}")
        PDF_LIBRARY = None


def generate_pdf_from_html(html_content, output_path=None):
    try:
        if PDF_LIBRARY == 'weasyprint' and IS_RENDER:
            from weasyprint import HTML
            if output_path:
                HTML(string=html_content).write_pdf(output_path)
                return output_path
            else:
                return HTML(string=html_content).write_pdf()
        elif PDF_LIBRARY == 'pdfkit':
            import pdfkit
            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                f.write(html_content)
                temp_html = f.name

            if output_path:
                pdfkit.from_file(temp_html, output_path, configuration=pdf_config, options={
                    'page-size': 'A4',
                    'encoding': 'UTF-8',
                    'margin-top': '10mm',
                    'margin-right': '10mm',
                    'margin-bottom': '10mm',
                    'margin-left': '10mm'
                })
                result = output_path
            else:
                temp_pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
                pdf_path = temp_pdf.name
                temp_pdf.close()
                pdfkit.from_file(temp_html, pdf_path, configuration=pdf_config, options={
                    'page-size': 'A4',
                    'encoding': 'UTF-8',
                    'margin-top': '10mm',
                    'margin-right': '10mm',
                    'margin-bottom': '10mm',
                    'margin-left': '10mm'
                })
                with open(pdf_path, 'rb') as f:
                    pdf_data = f.read()
                os.unlink(pdf_path)
                result = pdf_data
            os.unlink(temp_html)
            return result
        else:
            raise Exception(f"No PDF library available")
    except Exception as e:
        print(f"PDF generation error: {e}")
        raise


def get_company_settings():
    if DATABASE_URL:
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM company_settings LIMIT 1")
        result = cursor.fetchone()
        conn.close()
        if result:
            columns = ['id', 'company_name', 'company_address', 'company_phone', 'company_email',
                       'company_website', 'company_registration', 'currency_symbol', 'currency_code',
                       'tax_rate', 'created_date', 'updated_date']
            return dict(zip(columns, result))
    else:
        import sqlite3
        conn = sqlite3.connect('payment_system.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM company_settings LIMIT 1")
        result = cursor.fetchone()
        conn.close()
        if result:
            return dict(result)

    return {
        'company_name': 'My Company',
        'company_address': '123 Business Street, City, Country',
        'company_phone': '+1 234 567 8900',
        'company_email': 'info@mycompany.com',
        'company_website': 'www.mycompany.com',
        'company_registration': 'REG-2024-001',
        'currency_symbol': '$',
        'currency_code': 'USD',
        'tax_rate': 0.0
    }


def get_company_logo():
    logo_path = os.path.join(app.config['LOGO_UPLOAD_FOLDER'], 'company_logo.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_data = base64.b64encode(f.read()).decode('utf-8')
            return f"data:image/png;base64,{logo_data}"
    return None


def amount_in_words(amount):
    def number_to_words(n):
        ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine']
        teens = ['Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
        tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']

        if n < 10:
            return ones[int(n)]
        elif n < 20:
            return teens[int(n) - 10]
        elif n < 100:
            return tens[int(n // 10)] + (' ' + ones[int(n % 10)] if n % 10 != 0 else '')
        elif n < 1000:
            return ones[int(n // 100)] + ' Hundred' + (' ' + number_to_words(n % 100) if n % 100 != 0 else '')
        elif n < 1000000:
            return number_to_words(int(n // 1000)) + ' Thousand' + (' ' + number_to_words(n % 1000) if n % 1000 != 0 else '')
        elif n < 1000000000:
            return number_to_words(int(n // 1000000)) + ' Million' + (' ' + number_to_words(n % 1000000) if n % 1000000 != 0 else '')
        else:
            return str(n)

    dollars = int(amount)
    cents = int(round((amount - dollars) * 100))

    if dollars == 0 and cents == 0:
        return "Zero"

    words = number_to_words(dollars)
    if cents > 0:
        words += f" and {cents:02d}/100"

    return words


@app.context_processor
def inject_company_settings():
    settings = get_company_settings()
    logo = get_company_logo()
    return {
        'company_settings': settings,
        'company_logo': logo
    }


app.jinja_env.globals.update(amount_in_words=amount_in_words)


# ============ ROUTES ============

@app.route('/')
def index():
    db = get_db()
    if DATABASE_URL:
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM suppliers")
        suppliers_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM payments")
        payments_count = cursor.fetchone()[0]
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_amount = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COALESCE(SUM(total), 0) FROM services WHERE status = 'Pending'")
        pending_amount = cursor.fetchone()[0] or 0
    else:
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM suppliers")
        suppliers_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM payments")
        payments_count = cursor.fetchone()[0]
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_amount = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COALESCE(SUM(total), 0) FROM services WHERE status = 'Pending'")
        pending_amount = cursor.fetchone()[0] or 0
    db.close()

    return render_template('index.html',
                           suppliers_count=suppliers_count,
                           payments_count=payments_count,
                           total_amount=f"{total_amount:,.0f}",
                           pending_amount=f"{pending_amount:,.0f}",
                           now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))


@app.route('/suppliers', methods=['GET', 'POST'])
def suppliers():
    db = get_db()
    cursor = db.cursor()

    if request.method == 'POST':
        try:
            name = request.form['name']
            role = request.form.get('role', '')
            address = request.form.get('address', '')
            telephone = request.form.get('telephone', '')
            bank_name = request.form.get('bank_name', '')
            account_number = request.form.get('account_number', '')
            account_name = request.form.get('account_name', '')
            swift_code = request.form.get('swift_code', '')
            bank_address = request.form.get('bank_address', '')
            id_type = request.form.get('id_type', '')
            id_number = request.form.get('id_number', '')
            street_address = request.form.get('street_address', '')

            if DATABASE_URL:
                cursor.execute('''INSERT INTO suppliers 
                    (name, role, address, telephone, bank_name, account_number, 
                     account_name, swift_code, bank_address, id_type, 
                     id_number, street_address)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                    (name, role, address, telephone, bank_name, account_number,
                     account_name, swift_code, bank_address, id_type,
                     id_number, street_address))
            else:
                cursor.execute('''INSERT INTO suppliers 
                    (name, role, address, telephone, bank_name, account_number, 
                     account_name, swift_code, bank_address, id_type, 
                     id_number, street_address)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (name, role, address, telephone, bank_name, account_number,
                     account_name, swift_code, bank_address, id_type,
                     id_number, street_address))

            db.commit()
            flash('Supplier added successfully!', 'success')
        except Exception as e:
            db.rollback()
            flash(f'Error: {str(e)}', 'error')
        finally:
            db.close()
        return redirect(url_for('suppliers'))

    if DATABASE_URL:
        cursor.execute("SELECT * FROM suppliers ORDER BY id DESC")
        suppliers_list = cursor.fetchall()
    else:
        cursor.execute("SELECT * FROM suppliers ORDER BY id DESC")
        suppliers_list = cursor.fetchall()
    db.close()
    return render_template('suppliers.html', suppliers=suppliers_list)


@app.route('/update_supplier', methods=['POST'])
def update_supplier():
    db = get_db()
    cursor = db.cursor()

    try:
        supplier_id = request.form['id']
        name = request.form['name']
        role = request.form.get('role', '')
        address = request.form.get('address', '')
        telephone = request.form.get('telephone', '')
        bank_name = request.form.get('bank_name', '')
        account_number = request.form.get('account_number', '')
        account_name = request.form.get('account_name', '')
        swift_code = request.form.get('swift_code', '')
        bank_address = request.form.get('bank_address', '')
        id_type = request.form.get('id_type', '')
        id_number = request.form.get('id_number', '')
        street_address = request.form.get('street_address', '')

        if DATABASE_URL:
            cursor.execute('''UPDATE suppliers SET 
                name = %s, role = %s, address = %s, telephone = %s,
                bank_name = %s, account_number = %s, account_name = %s,
                swift_code = %s, bank_address = %s, id_type = %s,
                id_number = %s, street_address = %s
                WHERE id = %s''',
                (name, role, address, telephone, bank_name, account_number,
                 account_name, swift_code, bank_address, id_type,
                 id_number, street_address, supplier_id))
        else:
            cursor.execute('''UPDATE suppliers SET 
                name = ?, role = ?, address = ?, telephone = ?,
                bank_name = ?, account_number = ?, account_name = ?,
                swift_code = ?, bank_address = ?, id_type = ?,
                id_number = ?, street_address = ?
                WHERE id = ?''',
                (name, role, address, telephone, bank_name, account_number,
                 account_name, swift_code, bank_address, id_type,
                 id_number, street_address, supplier_id))

        db.commit()
        flash('Supplier updated successfully!', 'success')
    except Exception as e:
        db.rollback()
        flash(f'Error: {str(e)}', 'error')
    finally:
        db.close()
    return redirect(url_for('suppliers'))


@app.route('/delete_supplier/<int:id>', methods=['POST'])
def delete_supplier(id):
    db = get_db()
    cursor = db.cursor()
    try:
        if DATABASE_URL:
            cursor.execute("DELETE FROM payments WHERE supplier_id = %s", (id,))
            cursor.execute("DELETE FROM services WHERE supplier_id = %s", (id,))
            cursor.execute("DELETE FROM suppliers WHERE id = %s", (id,))
        else:
            cursor.execute("DELETE FROM payments WHERE supplier_id = ?", (id,))
            cursor.execute("DELETE FROM services WHERE supplier_id = ?", (id,))
            cursor.execute("DELETE FROM suppliers WHERE id = ?", (id,))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': str(e)})
    finally:
        db.close()


@app.route('/supplier/<int:id>')
def view_supplier(id):
    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute("SELECT * FROM suppliers WHERE id = %s", (id,))
    else:
        cursor.execute("SELECT * FROM suppliers WHERE id = ?", (id,))
    supplier = cursor.fetchone()
    db.close()
    if DATABASE_URL:
        columns = ['id', 'name', 'role', 'address', 'telephone', 'bank_name',
                   'account_number', 'account_name', 'swift_code', 'bank_address',
                   'id_type', 'id_number', 'street_address', 'created_date']
        return jsonify(dict(zip(columns, supplier)))
    return jsonify(dict(supplier))


@app.route('/payments')
def payments():
    db = get_db()
    cursor = db.cursor()

    if DATABASE_URL:
        cursor.execute("SELECT * FROM suppliers ORDER BY name")
        suppliers = cursor.fetchall()
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, sv.service_description 
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            ORDER BY p.payment_date DESC
        ''')
        payments = cursor.fetchall()
        cursor.execute('''
            SELECT s.*, sup.name as supplier_name 
            FROM services s
            JOIN suppliers sup ON s.supplier_id = sup.id
            WHERE s.status = 'Pending'
            ORDER BY s.created_date DESC
        ''')
        pending_services = cursor.fetchall()
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_paid = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COALESCE(SUM(total), 0) FROM services WHERE status = 'Pending'")
        pending_amount = cursor.fetchone()[0] or 0
    else:
        cursor.execute("SELECT * FROM suppliers ORDER BY name")
        suppliers = cursor.fetchall()
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, sv.service_description 
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            ORDER BY p.payment_date DESC
        ''')
        payments = cursor.fetchall()
        cursor.execute('''
            SELECT s.*, sup.name as supplier_name 
            FROM services s
            JOIN suppliers sup ON s.supplier_id = sup.id
            WHERE s.status = 'Pending'
            ORDER BY s.created_date DESC
        ''')
        pending_services = cursor.fetchall()
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_paid = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COALESCE(SUM(total), 0) FROM services WHERE status = 'Pending'")
        pending_amount = cursor.fetchone()[0] or 0

    db.close()
    return render_template('payments.html',
                           suppliers=suppliers,
                           payments=payments,
                           pending_services=pending_services,
                           total_paid=total_paid,
                           pending_amount=pending_amount,
                           now=date.today().isoformat())


@app.route('/process_multiple_payments', methods=['POST'])
def process_multiple_payments():
    db = get_db()
    cursor = db.cursor()
    successful = 0
    failed = 0

    try:
        supplier_ids = request.form.getlist('supplier_id[]')
        service_ids = request.form.getlist('service_id[]')
        amounts = request.form.getlist('amount[]')
        payment_methods = request.form.getlist('payment_method[]')
        reference_numbers = request.form.getlist('reference_number[]')
        payment_dates = request.form.getlist('payment_date[]')
        notes_list = request.form.getlist('notes[]')
        service_descriptions = request.form.getlist('service_description[]')
        payment_types = request.form.getlist('payment_type[]')

        for i in range(len(supplier_ids)):
            try:
                if not supplier_ids[i] or not payment_methods[i]:
                    failed += 1
                    continue

                service_id = service_ids[i] if i < len(service_ids) and service_ids[i] else None
                amount = float(amounts[i]) if i < len(amounts) else 0
                service_description = service_descriptions[i] if i < len(service_descriptions) else ''
                payment_type = payment_types[i] if i < len(payment_types) else 'Earning'

                if not service_id:
                    if DATABASE_URL:
                        cursor.execute('''INSERT INTO services (supplier_id, service_description, total, status)
                            VALUES (%s, %s, %s, %s) RETURNING id''',
                            (supplier_ids[i], service_description, amount, 'Pending'))
                        service_id = cursor.fetchone()[0]
                    else:
                        cursor.execute('''INSERT INTO services (supplier_id, service_description, total, status)
                            VALUES (?, ?, ?, ?)''',
                            (supplier_ids[i], service_description, amount, 'Pending'))
                        service_id = cursor.lastrowid

                payment_date = payment_dates[i] if i < len(payment_dates) and payment_dates[i] else date.today().isoformat()
                reference_number = reference_numbers[i] if i < len(reference_numbers) else ''
                notes = notes_list[i] if i < len(notes_list) else ''

                if DATABASE_URL:
                    cursor.execute('''INSERT INTO payments 
                        (supplier_id, service_id, amount, payment_method, 
                         reference_number, payment_date, notes, payment_type)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                        (supplier_ids[i], service_id, amount, payment_methods[i],
                         reference_number, payment_date, notes, payment_type))
                else:
                    cursor.execute('''INSERT INTO payments 
                        (supplier_id, service_id, amount, payment_method, 
                         reference_number, payment_date, notes, payment_type)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                        (supplier_ids[i], service_id, amount, payment_methods[i],
                         reference_number, payment_date, notes, payment_type))

                if DATABASE_URL:
                    cursor.execute("UPDATE services SET status = %s WHERE id = %s", ('Paid', service_id))
                else:
                    cursor.execute("UPDATE services SET status = ? WHERE id = ?", ('Paid', service_id))
                successful += 1

            except Exception as e:
                print(f"Error: {e}")
                failed += 1
                continue

        db.commit()
        flash(f'Successfully processed {successful} payment(s)! {failed} failed.', 'success')
    except Exception as e:
        db.rollback()
        flash(f'Error: {str(e)}', 'error')
    finally:
        db.close()
    return redirect(url_for('payments'))


@app.route('/get_supplier_services/<int:supplier_id>')
def get_supplier_services(supplier_id):
    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute('''
            SELECT * FROM services 
            WHERE supplier_id = %s AND status = 'Pending'
            ORDER BY created_date DESC
        ''', (supplier_id,))
    else:
        cursor.execute('''
            SELECT * FROM services 
            WHERE supplier_id = ? AND status = 'Pending'
            ORDER BY created_date DESC
        ''', (supplier_id,))
    services = cursor.fetchall()
    db.close()

    services_list = []
    for service in services:
        if DATABASE_URL:
            columns = ['id', 'supplier_id', 'service_description', 'total', 'status', 'created_date']
            service_dict = dict(zip(columns, service))
        else:
            service_dict = dict(service)
        services_list.append({
            'id': service_dict['id'],
            'service_description': service_dict['service_description'],
            'amount': service_dict['total']
        })

    return jsonify({'services': services_list})


@app.route('/payment_history')
def payment_history():
    db = get_db()
    cursor = db.cursor()

    if DATABASE_URL:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, sv.service_description, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            ORDER BY p.payment_date DESC
        ''')
    else:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, sv.service_description, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            ORDER BY p.payment_date DESC
        ''')
    payments_data = cursor.fetchall()
    db.close()

    payments = []
    for payment in payments_data:
        if DATABASE_URL:
            columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                       'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                       'created_date', 'supplier_name', 'service_description']
            payment_dict = dict(zip(columns, payment))
        else:
            payment_dict = dict(payment)
        payments.append({
            'id': payment_dict['id'],
            'payment_date': payment_dict['payment_date'],
            'supplier_name': payment_dict['supplier_name'],
            'service_description': payment_dict['service_description'],
            'amount': payment_dict['amount'],
            'payment_method': payment_dict['payment_method'],
            'reference_number': payment_dict.get('reference_number', ''),
            'notes': payment_dict.get('notes', ''),
            'payment_type': payment_dict.get('payment_type', 'Earning'),
            'supplier_id': payment_dict['supplier_id']
        })

    suppliers = []
    if DATABASE_URL:
        cursor = db.cursor()
        cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
        suppliers_data = cursor.fetchall()
        db.close()
        for s in suppliers_data:
            suppliers.append({'id': s[0], 'name': s[1]})
    else:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
        suppliers_data = cursor.fetchall()
        db.close()
        for s in suppliers_data:
            suppliers.append({'id': s[0], 'name': s[1]})

    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_paid = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments WHERE strftime('%Y-%m', payment_date) = strftime('%Y-%m', 'now')")
        monthly_total = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(DISTINCT supplier_id) FROM payments")
        unique_suppliers = cursor.fetchone()[0] or 0
    else:
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_paid = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments WHERE strftime('%Y-%m', payment_date) = strftime('%Y-%m', 'now')")
        monthly_total = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(DISTINCT supplier_id) FROM payments")
        unique_suppliers = cursor.fetchone()[0] or 0
    db.close()

    return render_template('payment_history.html',
                           payments=payments,
                           suppliers=suppliers,
                           total_paid=total_paid,
                           monthly_total=monthly_total,
                           unique_suppliers=unique_suppliers)


@app.route('/invoice/<int:payment_id>')
def invoice(payment_id):
    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number, s.swift_code,
                   s.bank_address, s.id_type, s.id_number, s.street_address,
                   sv.service_description, sv.total
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = %s
        ''', (payment_id,))
    else:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number, s.swift_code,
                   s.bank_address, s.id_type, s.id_number, s.street_address,
                   sv.service_description, sv.total
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = ?
        ''', (payment_id,))
    payment = cursor.fetchone()
    db.close()

    if not payment:
        flash('Payment not found', 'error')
        return redirect(url_for('payment_history'))

    if DATABASE_URL:
        columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                   'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                   'created_date', 'supplier_name', 'address', 'telephone', 'bank_name',
                   'account_name', 'account_number', 'swift_code', 'bank_address',
                   'id_type', 'id_number', 'street_address', 'service_description', 'total']
        payment_dict = dict(zip(columns, payment))
    else:
        payment_dict = dict(payment)

    return render_template('invoice.html', payment=payment_dict, payment_id=payment_id)


@app.route('/export_invoice_pdf/<int:payment_id>')
def export_invoice_pdf(payment_id):
    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number, s.swift_code,
                   s.bank_address, s.id_type, s.id_number, s.street_address,
                   sv.service_description, sv.total
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = %s
        ''', (payment_id,))
    else:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number, s.swift_code,
                   s.bank_address, s.id_type, s.id_number, s.street_address,
                   sv.service_description, sv.total
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = ?
        ''', (payment_id,))
    payment = cursor.fetchone()
    db.close()

    if not payment:
        flash('Payment not found', 'error')
        return redirect(url_for('payment_history'))

    if DATABASE_URL:
        columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                   'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                   'created_date', 'supplier_name', 'address', 'telephone', 'bank_name',
                   'account_name', 'account_number', 'swift_code', 'bank_address',
                   'id_type', 'id_number', 'street_address', 'service_description', 'total']
        payment_dict = dict(zip(columns, payment))
    else:
        payment_dict = dict(payment)

    html_content = render_template('invoice_pdf.html', payment=payment_dict, payment_id=payment_id)

    try:
        pdf_data = generate_pdf_from_html(html_content)
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=invoice_{payment_id}.pdf'
        return response
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('payment_history'))


@app.route('/bulk_invoice_pdf', methods=['POST'])
def bulk_invoice_pdf():
    payment_ids = request.form.getlist('payment_ids[]')

    if not payment_ids:
        flash('No payments selected', 'error')
        return redirect(url_for('payment_history'))

    db = get_db()
    cursor = db.cursor()
    payments = []

    for pid in payment_ids:
        if DATABASE_URL:
            cursor.execute('''
                SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                       s.bank_name, s.account_name, s.account_number, s.swift_code,
                       s.bank_address, s.id_type, s.id_number, s.street_address,
                       sv.service_description, sv.total
                FROM payments p
                JOIN suppliers s ON p.supplier_id = s.id
                JOIN services sv ON p.service_id = sv.id
                WHERE p.id = %s
            ''', (pid,))
        else:
            cursor.execute('''
                SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                       s.bank_name, s.account_name, s.account_number, s.swift_code,
                       s.bank_address, s.id_type, s.id_number, s.street_address,
                       sv.service_description, sv.total
                FROM payments p
                JOIN suppliers s ON p.supplier_id = s.id
                JOIN services sv ON p.service_id = sv.id
                WHERE p.id = ?
            ''', (pid,))
        payment = cursor.fetchone()
        if payment:
            if DATABASE_URL:
                columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                           'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                           'created_date', 'supplier_name', 'address', 'telephone', 'bank_name',
                           'account_name', 'account_number', 'swift_code', 'bank_address',
                           'id_type', 'id_number', 'street_address', 'service_description', 'total']
                payments.append(dict(zip(columns, payment)))
            else:
                payments.append(dict(payment))
    db.close()

    if not payments:
        flash('No invoices found', 'error')
        return redirect(url_for('payment_history'))

    try:
        temp_dir = tempfile.mkdtemp()
        pdf_files = []

        for payment in payments:
            html_content = render_template('invoice_pdf.html', payment=payment, payment_id=payment['id'])
            pdf_filename = f"invoice_{payment['id']}_{payment['supplier_name'].replace(' ', '_')}.pdf"
            pdf_path = os.path.join(temp_dir, pdf_filename)
            generate_pdf_from_html(html_content, pdf_path)
            pdf_files.append(pdf_path)

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for pdf_path in pdf_files:
                zip_file.write(pdf_path, os.path.basename(pdf_path))
                os.unlink(pdf_path)

        os.rmdir(temp_dir)
        zip_buffer.seek(0)
        response = make_response(zip_buffer.getvalue())
        response.headers['Content-Type'] = 'application/zip'
        response.headers['Content-Disposition'] = f'attachment; filename=invoices_{datetime.now().strftime("%Y%m%d")}.zip'
        return response

    except Exception as e:
        flash(f'Error generating invoices: {str(e)}', 'error')
        return redirect(url_for('payment_history'))


@app.route('/payslip/<int:payment_id>')
def payslip(payment_id):
    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number, s.swift_code,
                   s.bank_address, s.id_type, s.id_number, s.street_address,
                   s.role, sv.service_description, sv.total, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = %s
        ''', (payment_id,))
    else:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number, s.swift_code,
                   s.bank_address, s.id_type, s.id_number, s.street_address,
                   s.role, sv.service_description, sv.total, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = ?
        ''', (payment_id,))
    payment = cursor.fetchone()
    db.close()

    if not payment:
        flash('Payment not found', 'error')
        return redirect(url_for('payment_history'))

    if DATABASE_URL:
        columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                   'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                   'created_date', 'supplier_name', 'address', 'telephone', 'bank_name',
                   'account_name', 'account_number', 'swift_code', 'bank_address',
                   'id_type', 'id_number', 'street_address', 'role', 'service_description', 'total']
        payment_dict = dict(zip(columns, payment))
    else:
        payment_dict = dict(payment)

    payment_period = 'N/A'
    if payment_dict.get('payment_date'):
        try:
            payment_date = datetime.strptime(str(payment_dict['payment_date']), '%Y-%m-%d')
            payment_period = payment_date.strftime('%B %Y')
        except:
            payment_period = str(payment_dict['payment_date'])[:7]

    return render_template('payslip_report.html',
                           payment=payment_dict,
                           payment_id=payment_id,
                           now=datetime.now(),
                           payment_period=payment_period)


@app.route('/payslip/pdf/<int:payment_id>')
def payslip_pdf(payment_id):
    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number, s.swift_code,
                   s.bank_address, s.id_type, s.id_number, s.street_address,
                   s.role, sv.service_description, sv.total, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = %s
        ''', (payment_id,))
    else:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number, s.swift_code,
                   s.bank_address, s.id_type, s.id_number, s.street_address,
                   s.role, sv.service_description, sv.total, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = ?
        ''', (payment_id,))
    payment = cursor.fetchone()
    db.close()

    if not payment:
        flash('Payment not found', 'error')
        return redirect(url_for('payment_history'))

    if DATABASE_URL:
        columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                   'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                   'created_date', 'supplier_name', 'address', 'telephone', 'bank_name',
                   'account_name', 'account_number', 'swift_code', 'bank_address',
                   'id_type', 'id_number', 'street_address', 'role', 'service_description', 'total']
        payment_dict = dict(zip(columns, payment))
    else:
        payment_dict = dict(payment)

    payment_period = 'N/A'
    if payment_dict.get('payment_date'):
        try:
            payment_date = datetime.strptime(str(payment_dict['payment_date']), '%Y-%m-%d')
            payment_period = payment_date.strftime('%B %Y')
        except:
            payment_period = str(payment_dict['payment_date'])[:7]

    html_content = render_template('payslip_pdf.html',
                                   payment=payment_dict,
                                   payment_id=payment_id,
                                   now=datetime.now(),
                                   payment_period=payment_period)

    try:
        pdf_data = generate_pdf_from_html(html_content)
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=payslip_{payment_id}.pdf'
        return response
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('payslip', payment_id=payment_id))


@app.route('/payslip/list')
def payslip_list():
    db = get_db()
    cursor = db.cursor()

    if DATABASE_URL:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, sv.service_description, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            ORDER BY p.payment_date DESC
        ''')
    else:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, sv.service_description, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            ORDER BY p.payment_date DESC
        ''')
    payments_data = cursor.fetchall()
    db.close()

    payments = []
    pay_periods_set = set()

    for payment in payments_data:
        if DATABASE_URL:
            columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                       'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                       'created_date', 'supplier_name', 'service_description']
            payment_dict = dict(zip(columns, payment))
        else:
            payment_dict = dict(payment)

        pay_period = 'N/A'
        if payment_dict.get('payment_date'):
            try:
                payment_date = datetime.strptime(str(payment_dict['payment_date']), '%Y-%m-%d')
                pay_period = payment_date.strftime('%B %Y')
            except:
                pay_period = str(payment_dict['payment_date'])[:7]

        pay_periods_set.add(pay_period)

        payments.append({
            'id': payment_dict['id'],
            'payment_date': payment_dict['payment_date'],
            'pay_period': pay_period,
            'supplier_name': payment_dict['supplier_name'],
            'service_description': payment_dict['service_description'],
            'amount': payment_dict['amount'],
            'payment_method': payment_dict['payment_method'],
            'payment_type': payment_dict['payment_type'] or 'Earning',
            'supplier_id': payment_dict['supplier_id']
        })

    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
        suppliers_data = cursor.fetchall()
        db.close()
        suppliers = [{'id': s[0], 'name': s[1]} for s in suppliers_data]
    else:
        cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
        suppliers_data = cursor.fetchall()
        db.close()
        suppliers = [{'id': s[0], 'name': s[1]} for s in suppliers_data]

    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_amount = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments WHERE strftime('%Y-%m', payment_date) = strftime('%Y-%m', 'now')")
        monthly_total = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(DISTINCT supplier_id) FROM payments")
        unique_suppliers = cursor.fetchone()[0] or 0
    else:
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_amount = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments WHERE strftime('%Y-%m', payment_date) = strftime('%Y-%m', 'now')")
        monthly_total = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(DISTINCT supplier_id) FROM payments")
        unique_suppliers = cursor.fetchone()[0] or 0
    db.close()

    def parse_pay_period(period):
        if period == 'N/A':
            return datetime(1900, 1, 1)
        try:
            return datetime.strptime(period, '%B %Y')
        except:
            return datetime(1900, 1, 1)

    pay_periods = sorted(list(pay_periods_set), key=parse_pay_period, reverse=True)

    return render_template('payslip_list.html',
                           payments=payments,
                           pay_periods=pay_periods,
                           suppliers=suppliers,
                           total_amount=total_amount,
                           monthly_total=monthly_total,
                           unique_suppliers=unique_suppliers)


@app.route('/bulk_payslip_pdf', methods=['POST'])
def bulk_payslip_pdf():
    payment_ids = request.form.getlist('payment_ids[]')

    if not payment_ids:
        flash('No payslips selected', 'error')
        return redirect(url_for('payslip_list'))

    db = get_db()
    cursor = db.cursor()
    payments = []

    for pid in payment_ids:
        if DATABASE_URL:
            cursor.execute('''
                SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                       s.bank_name, s.account_name, s.account_number, s.swift_code,
                       s.bank_address, s.id_type, s.id_number, s.street_address,
                       s.role, sv.service_description, sv.total, p.payment_type
                FROM payments p
                JOIN suppliers s ON p.supplier_id = s.id
                JOIN services sv ON p.service_id = sv.id
                WHERE p.id = %s
            ''', (pid,))
        else:
            cursor.execute('''
                SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                       s.bank_name, s.account_name, s.account_number, s.swift_code,
                       s.bank_address, s.id_type, s.id_number, s.street_address,
                       s.role, sv.service_description, sv.total, p.payment_type
                FROM payments p
                JOIN suppliers s ON p.supplier_id = s.id
                JOIN services sv ON p.service_id = sv.id
                WHERE p.id = ?
            ''', (pid,))
        payment = cursor.fetchone()
        if payment:
            if DATABASE_URL:
                columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                           'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                           'created_date', 'supplier_name', 'address', 'telephone', 'bank_name',
                           'account_name', 'account_number', 'swift_code', 'bank_address',
                           'id_type', 'id_number', 'street_address', 'role', 'service_description', 'total']
                payments.append(dict(zip(columns, payment)))
            else:
                payments.append(dict(payment))
    db.close()

    if not payments:
        flash('No payslips found', 'error')
        return redirect(url_for('payslip_list'))

    try:
        temp_dir = tempfile.mkdtemp()
        pdf_files = []

        for payment in payments:
            payment_period = 'N/A'
            if payment.get('payment_date'):
                try:
                    payment_date = datetime.strptime(str(payment['payment_date']), '%Y-%m-%d')
                    payment_period = payment_date.strftime('%B %Y')
                except:
                    payment_period = str(payment['payment_date'])[:7]

            html_content = render_template('payslip_pdf.html',
                                         payment=payment,
                                         payment_id=payment['id'],
                                         now=datetime.now(),
                                         payment_period=payment_period)

            pdf_filename = f"payslip_{payment['id']}_{payment['supplier_name'].replace(' ', '_')}.pdf"
            pdf_path = os.path.join(temp_dir, pdf_filename)
            generate_pdf_from_html(html_content, pdf_path)
            pdf_files.append(pdf_path)

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for pdf_path in pdf_files:
                zip_file.write(pdf_path, os.path.basename(pdf_path))
                os.unlink(pdf_path)

        os.rmdir(temp_dir)
        zip_buffer.seek(0)
        response = make_response(zip_buffer.getvalue())
        response.headers['Content-Type'] = 'application/zip'
        response.headers['Content-Disposition'] = f'attachment; filename=payslips_{datetime.now().strftime("%Y%m%d")}.zip'
        return response

    except Exception as e:
        flash(f'Error generating payslips: {str(e)}', 'error')
        return redirect(url_for('payslip_list'))


@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        try:
            company_name = request.form.get('company_name', '')
            company_address = request.form.get('company_address', '')
            company_phone = request.form.get('company_phone', '')
            company_email = request.form.get('company_email', '')
            company_website = request.form.get('company_website', '')
            company_registration = request.form.get('company_registration', '')
            currency_symbol = request.form.get('currency_symbol', '$')
            currency_code = request.form.get('currency_code', 'USD')
            tax_rate = float(request.form.get('tax_rate', 0))

            db = get_db()
            cursor = db.cursor()

            if DATABASE_URL:
                cursor.execute("SELECT id FROM company_settings LIMIT 1")
                existing = cursor.fetchone()
                if existing:
                    cursor.execute('''UPDATE company_settings SET 
                        company_name = %s, company_address = %s, company_phone = %s,
                        company_email = %s, company_website = %s, company_registration = %s,
                        currency_symbol = %s, currency_code = %s, tax_rate = %s,
                        updated_date = CURRENT_TIMESTAMP
                        WHERE id = %s''',
                        (company_name, company_address, company_phone, company_email,
                         company_website, company_registration, currency_symbol,
                         currency_code, tax_rate, existing[0]))
                else:
                    cursor.execute('''INSERT INTO company_settings 
                        (company_name, company_address, company_phone, company_email,
                         company_website, company_registration, currency_symbol,
                         currency_code, tax_rate)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                        (company_name, company_address, company_phone, company_email,
                         company_website, company_registration, currency_symbol,
                         currency_code, tax_rate))
            else:
                cursor.execute("SELECT id FROM company_settings LIMIT 1")
                existing = cursor.fetchone()
                if existing:
                    cursor.execute('''UPDATE company_settings SET 
                        company_name = ?, company_address = ?, company_phone = ?,
                        company_email = ?, company_website = ?, company_registration = ?,
                        currency_symbol = ?, currency_code = ?, tax_rate = ?,
                        updated_date = CURRENT_TIMESTAMP
                        WHERE id = ?''',
                        (company_name, company_address, company_phone, company_email,
                         company_website, company_registration, currency_symbol,
                         currency_code, tax_rate, existing[0]))
                else:
                    cursor.execute('''INSERT INTO company_settings 
                        (company_name, company_address, company_phone, company_email,
                         company_website, company_registration, currency_symbol,
                         currency_code, tax_rate)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (company_name, company_address, company_phone, company_email,
                         company_website, company_registration, currency_symbol,
                         currency_code, tax_rate))

            db.commit()
            db.close()
            flash('Company settings updated successfully!', 'success')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('settings'))

    settings = get_company_settings()
    logo_path = os.path.join(app.config['LOGO_UPLOAD_FOLDER'], 'company_logo.png')
    logo_exists = os.path.exists(logo_path)

    return render_template('settings.html',
                           settings=settings,
                           logo_exists=logo_exists)


@app.route('/settings/logo', methods=['POST'])
def upload_logo():
    if 'logo' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('settings'))

    file = request.files['logo']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('settings'))

    if not allowed_image_file(file.filename):
        flash('Invalid file format. Please upload PNG, JPG, JPEG, GIF, or SVG', 'error')
        return redirect(url_for('settings'))

    try:
        filename = 'company_logo.png'
        filepath = os.path.join(app.config['LOGO_UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            os.remove(filepath)
        file.save(filepath)
        flash('Company logo uploaded successfully!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('settings'))


@app.route('/settings/logo/remove', methods=['POST'])
def remove_logo():
    try:
        logo_path = os.path.join(app.config['LOGO_UPLOAD_FOLDER'], 'company_logo.png')
        if os.path.exists(logo_path):
            os.remove(logo_path)
            flash('Company logo removed successfully!', 'success')
        else:
            flash('No logo found to remove', 'warning')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('settings'))


@app.route('/export_all_payments_html')
def export_all_payments_html():
    db = get_db()
    cursor = db.cursor()

    if DATABASE_URL:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, sv.service_description, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            ORDER BY p.payment_date DESC
        ''')
    else:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, sv.service_description, p.payment_type
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            ORDER BY p.payment_date DESC
        ''')
    payments = cursor.fetchall()
    db.close()

    db = get_db()
    cursor = db.cursor()
    if DATABASE_URL:
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_amount = cursor.fetchone()[0] or 0
    else:
        cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM payments")
        total_amount = cursor.fetchone()[0] or 0
    db.close()

    # Convert payments to list of dicts for template
    payments_list = []
    for payment in payments:
        if DATABASE_URL:
            columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                       'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                       'created_date', 'supplier_name', 'service_description']
            payments_list.append(dict(zip(columns, payment)))
        else:
            payments_list.append(dict(payment))

    return render_template('all_payments_report.html',
                           payments=payments_list,
                           total_amount=total_amount,
                           now=datetime.now())


@app.route('/get_receipt/<int:payment_id>')
def get_receipt(payment_id):
    db = get_db()
    cursor = db.cursor()

    if DATABASE_URL:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number,
                   sv.service_description, sv.total
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = %s
        ''', (payment_id,))
    else:
        cursor.execute('''
            SELECT p.*, s.name as supplier_name, s.address, s.telephone,
                   s.bank_name, s.account_name, s.account_number,
                   sv.service_description, sv.total
            FROM payments p
            JOIN suppliers s ON p.supplier_id = s.id
            JOIN services sv ON p.service_id = sv.id
            WHERE p.id = ?
        ''', (payment_id,))
    payment = cursor.fetchone()
    db.close()

    if not payment:
        return jsonify({'success': False, 'error': 'Receipt not found'})

    if DATABASE_URL:
        columns = ['id', 'supplier_id', 'service_id', 'amount', 'payment_method',
                   'reference_number', 'payment_date', 'notes', 'payment_type', 'status',
                   'created_date', 'supplier_name', 'address', 'telephone', 'bank_name',
                   'account_name', 'account_number', 'service_description', 'total']
        payment_dict = dict(zip(columns, payment))
    else:
        payment_dict = dict(payment)

    receipt_html = f'''
    <div class="receipt" style="font-family: Arial, sans-serif; padding: 20px;">
        <div style="text-align: center; border-bottom: 2px solid #333; margin-bottom: 20px;">
            <h2>PAYMENT RECEIPT</h2>
            <p>Receipt No: {payment_dict['id']:06d}</p>
            <p>Date: {payment_dict['payment_date']}</p>
        </div>

        <div style="margin-bottom: 20px;">
            <h4>Supplier Information</h4>
            <p><strong>Name:</strong> {payment_dict['supplier_name']}</p>
            <p><strong>Address:</strong> {payment_dict.get('address', 'N/A')}</p>
            <p><strong>Telephone:</strong> {payment_dict.get('telephone', 'N/A')}</p>
        </div>

        <div style="margin-bottom: 20px;">
            <h4>Payment Details</h4>
            <p><strong>Service:</strong> {payment_dict['service_description']}</p>
            <p><strong>Amount:</strong> ${payment_dict['amount']:,.2f}</p>
            <p><strong>Payment Method:</strong> {payment_dict['payment_method']}</p>
            <p><strong>Reference:</strong> {payment_dict.get('reference_number', 'N/A')}</p>
            <p><strong>Payment Type:</strong> {payment_dict.get('payment_type', 'Earning')}</p>
        </div>

        <div style="margin-bottom: 20px;">
            <h4>Bank Details</h4>
            <p><strong>Bank Name:</strong> {payment_dict.get('bank_name', 'N/A')}</p>
            <p><strong>Account Name:</strong> {payment_dict.get('account_name', 'N/A')}</p>
            <p><strong>Account Number:</strong> {payment_dict.get('account_number', 'N/A')}</p>
        </div>

        <div style="margin-top: 30px; text-align: center;">
            <p>Thank you for your payment!</p>
        </div>
    </div>
    '''

    return jsonify({'success': True, 'html': receipt_html})


@app.route('/delete_payment/<int:payment_id>', methods=['POST'])
def delete_payment(payment_id):
    db = get_db()
    cursor = db.cursor()

    try:
        if DATABASE_URL:
            cursor.execute("SELECT service_id FROM payments WHERE id = %s", (payment_id,))
        else:
            cursor.execute("SELECT service_id FROM payments WHERE id = ?", (payment_id,))
        payment = cursor.fetchone()

        if payment:
            if DATABASE_URL:
                cursor.execute("DELETE FROM payments WHERE id = %s", (payment_id,))
                cursor.execute("UPDATE services SET status = %s WHERE id = %s", ('Pending', payment[0]))
            else:
                cursor.execute("DELETE FROM payments WHERE id = ?", (payment_id,))
                cursor.execute("UPDATE services SET status = ? WHERE id = ?", ('Pending', payment[0]))
            db.commit()
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Payment not found'})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': str(e)})
    finally:
        db.close()


@app.route('/download_payment_template')
def download_payment_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Template"

    headers = ['Supplier Name', 'Service Description', 'Amount', 'Payment Method',
               'Reference Number', 'Payment Date', 'Payment Type', 'Notes']
    ws.append(headers)

    example_data = [
        ['ABC Company', 'Consulting Services', '500000', 'Bank Transfer',
         'TRX001', '2024-01-15', 'Earning', 'Monthly payment'],
        ['XYZ Limited', 'Web Development', '750000', 'Cash',
         '', '2024-01-20', 'Earning', 'Initial deposit'],
        ['ABC Company', 'Tax Deduction', '100000', 'Bank Transfer',
         'TAX001', '2024-01-15', 'Deduction', ''],
    ]

    for row in example_data:
        ws.append(row)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='payment_import_template.xlsx'
    )


@app.route('/import_payments_excel', methods=['POST'])
def import_payments_excel():
    if 'payment_file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('payments'))

    file = request.files['payment_file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('payments'))

    if not allowed_file(file.filename):
        flash('Invalid file format. Please upload .xlsx, .xls, or .csv file', 'error')
        return redirect(url_for('payments'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    db = get_db()
    cursor = db.cursor()
    successful = 0
    failed = 0
    errors = []

    try:
        if filename.endswith('.csv'):
            with open(filepath, 'r', encoding='utf-8-sig') as csvfile:
                reader = csv.DictReader(csvfile)
                data = list(reader)
        else:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))

        for idx, row in enumerate(data, start=2):
            try:
                supplier_name = str(row.get('Supplier Name', '') or '').strip()
                service_description = str(row.get('Service Description', '') or '').strip()
                amount = row.get('Amount', 0) or 0
                payment_method = str(row.get('Payment Method', '') or '').strip()
                reference_number = str(row.get('Reference Number', '') or '').strip()
                payment_date = str(row.get('Payment Date', '') or '').strip()
                payment_type = str(row.get('Payment Type', '') or 'Earning').strip()
                notes = str(row.get('Notes', '') or '').strip()

                if not supplier_name or not service_description or not amount or not payment_method:
                    errors.append(f"Row {idx}: Missing required fields")
                    failed += 1
                    continue

                if DATABASE_URL:
                    cursor.execute("SELECT id FROM suppliers WHERE LOWER(name) LIKE LOWER(%s)", (f'%{supplier_name}%',))
                else:
                    cursor.execute("SELECT id FROM suppliers WHERE LOWER(name) LIKE LOWER(?)", (f'%{supplier_name}%',))
                supplier = cursor.fetchone()

                if not supplier:
                    errors.append(f"Row {idx}: Supplier '{supplier_name}' not found")
                    failed += 1
                    continue

                supplier_id = supplier[0]

                if DATABASE_URL:
                    cursor.execute('''INSERT INTO services (supplier_id, service_description, total, status)
                        VALUES (%s, %s, %s, %s) RETURNING id''',
                        (supplier_id, service_description, float(amount), 'Pending'))
                    service_id = cursor.fetchone()[0]
                else:
                    cursor.execute('''INSERT INTO services (supplier_id, service_description, total, status)
                        VALUES (?, ?, ?, ?)''',
                        (supplier_id, service_description, float(amount), 'Pending'))
                    service_id = cursor.lastrowid

                if not payment_date:
                    payment_date = date.today().isoformat()

                if DATABASE_URL:
                    cursor.execute('''INSERT INTO payments 
                        (supplier_id, service_id, amount, payment_method, 
                         reference_number, payment_date, notes, payment_type)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                        (supplier_id, service_id, float(amount), payment_method,
                         reference_number, payment_date, notes, payment_type))
                    cursor.execute("UPDATE services SET status = %s WHERE id = %s", ('Paid', service_id))
                else:
                    cursor.execute('''INSERT INTO payments 
                        (supplier_id, service_id, amount, payment_method, 
                         reference_number, payment_date, notes, payment_type)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                        (supplier_id, service_id, float(amount), payment_method,
                         reference_number, payment_date, notes, payment_type))
                    cursor.execute("UPDATE services SET status = ? WHERE id = ?", ('Paid', service_id))

                successful += 1

            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
                failed += 1
                continue

        db.commit()

        if successful > 0:
            flash(f'Successfully imported {successful} payment(s)!', 'success')
        if failed > 0:
            flash(f'{failed} payment(s) failed to import.', 'warning')
        if errors:
            error_msg = '; '.join(errors[:5])
            if len(errors) > 5:
                error_msg += f' and {len(errors) - 5} more errors'
            flash(f'Errors: {error_msg}', 'error')

    except Exception as e:
        db.rollback()
        flash(f'Error: {str(e)}', 'error')
    finally:
        db.close()
        time.sleep(0.5)
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except:
            pass

    return redirect(url_for('payments'))


@app.route('/bulk_payment_status')
def bulk_payment_status():
    db = get_db()
    cursor = db.cursor()

    if DATABASE_URL:
        cursor.execute('''
            SELECT COUNT(*) as count, COALESCE(SUM(amount), 0) as total
            FROM payments 
            WHERE created_date >= CURRENT_TIMESTAMP - INTERVAL '1 day'
        ''')
    else:
        cursor.execute('''
            SELECT COUNT(*) as count, COALESCE(SUM(amount), 0) as total
            FROM payments 
            WHERE created_date >= datetime('now', '-1 day')
        ''')
    result = cursor.fetchone()
    db.close()

    return jsonify({
        'count': result[0],
        'total': result[1] or 0
    })


@app.route('/export/suppliers/<format>')
def export_suppliers(format):
    """Export suppliers to CSV, Excel, or JSON"""
    db = get_db()
    cursor = db.cursor()

    if DATABASE_URL:
        cursor.execute("SELECT * FROM suppliers ORDER BY id DESC")
    else:
        cursor.execute("SELECT * FROM suppliers ORDER BY id DESC")
    suppliers = cursor.fetchall()
    db.close()

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'suppliers_{timestamp}'

    # Convert suppliers to list of dicts
    suppliers_list = []
    for supplier in suppliers:
        if DATABASE_URL:
            columns = ['id', 'name', 'role', 'address', 'telephone', 'bank_name',
                       'account_number', 'account_name', 'swift_code', 'bank_address',
                       'id_type', 'id_number', 'street_address', 'created_date']
            suppliers_list.append(dict(zip(columns, supplier)))
        else:
            suppliers_list.append(dict(supplier))

    if format == 'csv':
        output = StringIO()
        csv_writer = csv.writer(output)
        csv_writer.writerow(['ID', 'Name', 'Role', 'Address', 'Telephone', 'Bank Name',
                             'Account Number', 'Account Name', 'SWIFT Code',
                             'Bank Address', 'ID Type', 'ID Number',
                             'Street Address', 'Created Date'])

        for supplier in suppliers_list:
            csv_writer.writerow([
                supplier.get('id', ''),
                supplier.get('name', ''),
                supplier.get('role', ''),
                supplier.get('address', ''),
                supplier.get('telephone', ''),
                supplier.get('bank_name', ''),
                supplier.get('account_number', ''),
                supplier.get('account_name', ''),
                supplier.get('swift_code', ''),
                supplier.get('bank_address', ''),
                supplier.get('id_type', ''),
                supplier.get('id_number', ''),
                supplier.get('street_address', ''),
                supplier.get('created_date', '')
            ])

        output.seek(0)
        return send_file(
            BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'{filename}.csv'
        )

    elif format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Suppliers"

        headers = ['ID', 'Name', 'Role', 'Address', 'Telephone', 'Bank Name',
                   'Account Number', 'Account Name', 'SWIFT Code',
                   'Bank Address', 'ID Type', 'ID Number',
                   'Street Address', 'Created Date']
        ws.append(headers)

        for supplier in suppliers_list:
            ws.append([
                supplier.get('id', ''),
                supplier.get('name', ''),
                supplier.get('role', ''),
                supplier.get('address', ''),
                supplier.get('telephone', ''),
                supplier.get('bank_name', ''),
                supplier.get('account_number', ''),
                supplier.get('account_name', ''),
                supplier.get('swift_code', ''),
                supplier.get('bank_address', ''),
                supplier.get('id_type', ''),
                supplier.get('id_number', ''),
                supplier.get('street_address', ''),
                supplier.get('created_date', '')
            ])

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{filename}.xlsx'
        )

    elif format == 'json':
        return send_file(
            BytesIO(json.dumps(suppliers_list, indent=2, default=str).encode()),
            mimetype='application/json',
            as_attachment=True,
            download_name=f'{filename}.json'
        )

    return jsonify({'error': 'Invalid format'}), 400


@app.route('/import/suppliers', methods=['POST'])
def import_suppliers():
    """Import suppliers from CSV/Excel"""
    if 'file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('suppliers'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('suppliers'))

    if not file or not allowed_file(file.filename):
        flash('Invalid file format. Please upload .csv, .xlsx, or .xls file', 'error')
        return redirect(url_for('suppliers'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    db = get_db()
    cursor = db.cursor()
    success_count = 0
    error_count = 0

    try:
        if filename.endswith('.csv'):
            with open(filepath, 'r', encoding='utf-8-sig') as csvfile:
                csv_reader = csv.DictReader(csvfile)
                for row in csv_reader:
                    try:
                        if DATABASE_URL:
                            cursor.execute('''INSERT INTO suppliers 
                                (name, role, address, telephone, bank_name, account_number, 
                                 account_name, swift_code, bank_address, id_type, 
                                 id_number, street_address)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                                (row.get('Name', ''), row.get('Role', ''),
                                 row.get('Address', ''), row.get('Telephone', ''),
                                 row.get('Bank Name', ''), row.get('Account Number', ''),
                                 row.get('Account Name', ''), row.get('SWIFT Code', ''),
                                 row.get('Bank Address', ''), row.get('ID Type', ''),
                                 row.get('ID Number', ''), row.get('Street Address', '')))
                        else:
                            cursor.execute('''INSERT INTO suppliers 
                                (name, role, address, telephone, bank_name, account_number, 
                                 account_name, swift_code, bank_address, id_type, 
                                 id_number, street_address)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                                (row.get('Name', ''), row.get('Role', ''),
                                 row.get('Address', ''), row.get('Telephone', ''),
                                 row.get('Bank Name', ''), row.get('Account Number', ''),
                                 row.get('Account Name', ''), row.get('SWIFT Code', ''),
                                 row.get('Bank Address', ''), row.get('ID Type', ''),
                                 row.get('ID Number', ''), row.get('Street Address', '')))
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        continue

        elif filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    row_dict = dict(zip(headers, row))
                    if DATABASE_URL:
                        cursor.execute('''INSERT INTO suppliers 
                            (name, role, address, telephone, bank_name, account_number, 
                             account_name, swift_code, bank_address, id_type, 
                             id_number, street_address)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                            (row_dict.get('Name', ''), row_dict.get('Role', ''),
                             row_dict.get('Address', ''), row_dict.get('Telephone', ''),
                             row_dict.get('Bank Name', ''), row_dict.get('Account Number', ''),
                             row_dict.get('Account Name', ''), row_dict.get('SWIFT Code', ''),
                             row_dict.get('Bank Address', ''), row_dict.get('ID Type', ''),
                             row_dict.get('ID Number', ''), row_dict.get('Street Address', '')))
                    else:
                        cursor.execute('''INSERT INTO suppliers 
                            (name, role, address, telephone, bank_name, account_number, 
                             account_name, swift_code, bank_address, id_type, 
                             id_number, street_address)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                            (row_dict.get('Name', ''), row_dict.get('Role', ''),
                             row_dict.get('Address', ''), row_dict.get('Telephone', ''),
                             row_dict.get('Bank Name', ''), row_dict.get('Account Number', ''),
                             row_dict.get('Account Name', ''), row_dict.get('SWIFT Code', ''),
                             row_dict.get('Bank Address', ''), row_dict.get('ID Type', ''),
                             row_dict.get('ID Number', ''), row_dict.get('Street Address', '')))
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    continue

        db.commit()

        if success_count > 0:
            flash(f'Successfully imported {success_count} suppliers!', 'success')
        if error_count > 0:
            flash(f'{error_count} records failed to import.', 'warning')

    except Exception as e:
        db.rollback()
        flash(f'Error importing file: {str(e)}', 'error')

    finally:
        db.close()
        time.sleep(0.5)
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except:
            pass

    return redirect(url_for('suppliers'))

if __name__ == '__main__':
    app.run(debug=True)